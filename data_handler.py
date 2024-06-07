import openpyxl
from openpyxl import load_workbook
import warnings
import sqlite3
import numpy as np

warnings.filterwarnings("ignore")

# 엑셀 파일 열기
workbook = load_workbook(filename='./통합 식품영양성분DB_음식_20230715.xlsx')

# 첫 번째 시트 선택
sheet = workbook.active

# 데이터 읽기
raw_data = []
for row in sheet.iter_rows(values_only=True):
    raw_data.append(row)

if not raw_data:
    exit



# 데이터베이스 연결
conn = sqlite3.connect('food_nutrition.db')
# 커서 생성
cursor = conn.cursor()

#테이블 생성
column = ["id", "food_cd", "group_name", "food_name", "research_year", "maker_name", "ref_name", "serving_size", "calorie", "carbohydrate", "protein", "province", "sugers", "salt", "cholesterol", "saturated_fatty_acides", "trans_fat"]
create_table_query = f'''CREATE TABLE IF NOT EXISTS "data"(
                         "{column[0]}" INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
                         "{column[1]}" TEXT,
                         "{column[2]}" TEXT,
                         "{column[3]}" TEXT,
                         "{column[4]}" INTEGER,
                         "{column[5]}" TEXT,
                         "{column[6]}" TEXT,
                         "{column[7]}" INTERGER,
                         "{column[8]}" FLOAT,
                         "{column[9]}" FLOAT,
                         "{column[10]}" FLOAT,
                         "{column[11]}" FLOAT,
                         "{column[12]}" FLOAT,
                         "{column[13]}" FLOAT,
                         "{column[14]}" FLOAT,
                         "{column[15]}" FLOAT,
                         "{column[16]}" FLOAT)
                         '''

cursor.execute(create_table_query)
conn.commit()


# 컬럼 뽑기
comment = ["식품코드", "DB군", "식품명", "연도", "지역 / 제조사", "성분표출처", "1회제공량", "에너지(㎉)", "탄수화물(g)", "단백질(g)", "지방(g)", "총당류(g)", "나트륨(㎎)", "콜레스테롤(㎎)", "총 포화 지방산(g)", "트랜스 지방산(g)"]

index = []
for c in comment:
    for i in range(len(raw_data[0])) :
        if c in raw_data[0][i]:
            index.append(i)

b = np.array(raw_data[1:]).T
data = b[index]
data = np.array(data).T.tolist()

# 데이터 삽입
for row in data:
    data_insert_query = f"INSERT INTO data ({','.join(column[1:])}) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
    cursor.execute(data_insert_query, row)

#cursor.executemany("INSERT INTO data ({','.join(column[1:])}) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", data)

# 데이터베이스에 변경 사항 저장
conn.commit()


# 연결 종료
conn.close()
